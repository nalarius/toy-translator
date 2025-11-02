"""CLI for converting the toy translator dataset from XLSX to JSON."""

from __future__ import annotations

import argparse
import json
import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import google.generativeai as genai
from dotenv import load_dotenv


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert an XLSX dataset into JSON, omitting empty cells, and classify columns."
    )
    parser.add_argument(
        "--input",
        default=Path("data/source.xlsx"),
        type=Path,
        help="Path to the source XLSX file (default: data/source.xlsx).",
    )
    parser.add_argument(
        "--output",
        default=Path("tmp/source.json"),
        type=Path,
        help="Where to write the JSON output (default: tmp/source.json).",
    )
    parser.add_argument(
        "--schema-output",
        type=Path,
        default=Path("tmp/column_schema.json"),
        help="Where to save column classification schema (default: tmp/column_schema.json).",
    )
    parser.add_argument(
        "--model",
        default="gemini-2.5-flash",
        help="Gemini model for column classification (default: gemini-2.5-flash).",
    )
    parser.add_argument(
        "--sample-rows",
        type=int,
        default=10,
        help="Number of sample rows to send to Gemini for analysis (default: 10).",
    )
    parser.add_argument(
        "--skip-classification",
        action="store_true",
        help="Skip column classification and use default schema (KEY, Korean).",
    )
    return parser


def _normalise_value(value: Any) -> Any | None:
    """Return cleaned value or None (for empty)."""
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    return value


def configure_gemini(model_name: str) -> genai.GenerativeModel:
    """Configure Gemini API and return model instance."""
    load_dotenv()
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise EnvironmentError("GEMINI_API_KEY is not set in the environment or .env file.")
    genai.configure(api_key=api_key)
    return genai.GenerativeModel(model_name)


def extract_sample_data(data: list[dict[str, Any]], limit: int = 10) -> list[dict[str, Any]]:
    """Extract first N rows for Gemini analysis."""
    return data[:limit]


def build_column_classification_prompt(headers: list[str], sample_data: list[dict[str, Any]]) -> str:
    """Build prompt for Gemini to classify columns."""
    sample_json = json.dumps(sample_data, ensure_ascii=False, indent=2)
    headers_json = json.dumps(headers, ensure_ascii=False)

    return f"""
You are analyzing a game localization spreadsheet to classify columns for automated translation.

**Task**: Classify each column into one of these categories:

1. **unique_id** (exactly 1): The unique identifier for each row
   - Usually named: "id", "ID", "key", "KEY", "index", etc.
   - Must have unique values across all rows

2. **text_columns**: Text content to be translated
   - **primary** (exactly 1): Main text to translate (e.g., dialogue, description)
   - **secondary** (0+): Additional translatable text (e.g., alternative text, notes)
   - Usually contains natural language text, often in specific language (Korean, Japanese, etc.)

3. **metadata**: Additional information (categorize into subcategories):
   - **speaker**: Speaker/character identification (e.g., "speaker", "character", "NPC_name")
   - **context**: Contextual information (e.g., "scene", "location", "session_id", "quest")
   - **tags**: Categories/tags (e.g., "type", "category", "genre", "flag")
   - **other**: Any other metadata not fitting above categories

**Columns**:
{headers_json}

**Sample Data (first {len(sample_data)} rows)**:
{sample_json}

**Response Format** (JSON only, no markdown):
{{
  "unique_id": "column_name",
  "text_columns": {{
    "primary": "column_name",
    "secondary": ["column_name1", "column_name2"]
  }},
  "metadata": {{
    "speaker": ["column1", "column2"],
    "context": ["column3"],
    "tags": ["column4"],
    "other": ["column5", "column6"]
  }}
}}

**Rules**:
- Every column must be classified exactly once
- If uncertain, prefer metadata.other
- Consider both column names AND actual data values
- Empty/null values are normal in sample data
- Return valid JSON only, no additional text or markdown formatting
""".strip()


def call_gemini_for_classification(
    model: genai.GenerativeModel,
    headers: list[str],
    sample_data: list[dict[str, Any]]
) -> dict[str, Any]:
    """Call Gemini to classify columns and return parsed JSON."""
    prompt = build_column_classification_prompt(headers, sample_data)

    response = model.generate_content(
        prompt,
        generation_config=genai.GenerationConfig(
            temperature=0.2,
            response_mime_type="application/json",
        ),
    )

    if not response or not response.text:
        raise RuntimeError("Gemini returned an empty response.")

    # Parse JSON response
    text = response.text.strip()
    json_match = re.search(r"\{[\s\S]*\}", text)
    if not json_match:
        raise ValueError("Could not locate JSON object in Gemini response.")

    return json.loads(json_match.group(0))


def validate_uniqueness(data: list[dict[str, Any]], column: str) -> tuple[bool, list[Any]]:
    """
    Validate if column values are unique.
    Returns: (is_unique, list_of_duplicates)
    """
    values = [row.get(column) for row in data if row.get(column) is not None]
    seen: set[Any] = set()
    duplicates: list[Any] = []

    for value in values:
        if value in seen and value not in duplicates:
            duplicates.append(value)
        seen.add(value)

    return len(duplicates) == 0, duplicates


def create_default_schema(headers: list[str]) -> dict[str, Any]:
    """Create a default schema when classification is skipped."""
    # Try to find KEY and Korean columns
    unique_id = "KEY" if "KEY" in headers else headers[0]
    text_primary = "Korean" if "Korean" in headers else (headers[1] if len(headers) > 1 else headers[0])

    other_columns = [h for h in headers if h not in (unique_id, text_primary)]

    return {
        "unique_id": unique_id,
        "text_columns": {
            "primary": text_primary,
            "secondary": []
        },
        "metadata": {
            "speaker": [],
            "context": [],
            "tags": [],
            "other": other_columns
        }
    }


def save_column_schema(schema: dict[str, Any], output_path: Path) -> None:
    """Save column classification schema to JSON."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(schema, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def convert_xlsx_to_json(input_path: Path, output_path: Path) -> None:
    if not input_path.exists():
        raise FileNotFoundError(f"Input XLSX not found: {input_path}")

    workbook = load_workbook(filename=input_path, read_only=True)
    worksheet = workbook.active

    rows_iter = worksheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        raise ValueError("XLSX file does not contain a header row.")

    cleaned_headers = [
        header.strip() if isinstance(header, str) else header for header in headers
    ]

    result: list[dict[str, Any]] = []
    for row in rows_iter:
        row_dict: dict[str, Any] = {}
        for header, value in zip(cleaned_headers, row):
            if not header:
                continue
            normalised = _normalise_value(value)
            if normalised is None:
                continue
            row_dict[str(header)] = normalised
        if row_dict:
            result.append(row_dict)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    # Convert XLSX to JSON
    convert_xlsx_to_json(args.input, args.output)

    # Load the converted data for column classification
    dataset = json.loads(args.output.read_text(encoding="utf-8"))
    if not isinstance(dataset, list) or not dataset:
        print("Warning: Dataset is empty or not a list. Skipping column classification.")
        return

    headers = list(dataset[0].keys()) if dataset else []
    if not headers:
        print("Warning: No headers found in dataset. Skipping column classification.")
        return

    # Column classification
    if args.skip_classification:
        print("Skipping column classification (--skip-classification flag set).")
        schema = create_default_schema(headers)
        save_column_schema(schema, args.schema_output)
        print(f"Default schema saved to {args.schema_output}")
        return

    print(f"Analyzing {len(headers)} columns with Gemini...")
    print(f"Columns: {', '.join(headers)}")

    try:
        model = configure_gemini(args.model)
        sample_data = extract_sample_data(dataset, args.sample_rows)

        print(f"Sending {len(sample_data)} sample rows to Gemini for classification...")
        schema = call_gemini_for_classification(model, headers, sample_data)

        # Validate unique_id column
        unique_id_column = schema.get("unique_id")
        if unique_id_column:
            print(f"Validating uniqueness of '{unique_id_column}' column...")
            is_unique, duplicates = validate_uniqueness(dataset, unique_id_column)

            if is_unique:
                print(f"✓ Column '{unique_id_column}' has all unique values.")
            else:
                print(f"⚠ Warning: Column '{unique_id_column}' has {len(duplicates)} duplicate values:")
                for dup in duplicates[:5]:  # Show first 5 duplicates
                    print(f"  - {dup}")
                if len(duplicates) > 5:
                    print(f"  ... and {len(duplicates) - 5} more")

        # Save schema
        save_column_schema(schema, args.schema_output)
        print(f"✓ Column schema saved to {args.schema_output}")

        # Display classification summary
        print("\nColumn Classification Summary:")
        print(f"  Unique ID: {schema.get('unique_id')}")
        text_cols = schema.get('text_columns', {})
        print(f"  Primary Text: {text_cols.get('primary')}")
        if text_cols.get('secondary'):
            print(f"  Secondary Text: {', '.join(text_cols.get('secondary', []))}")

        metadata = schema.get('metadata', {})
        for category, columns in metadata.items():
            if columns:
                print(f"  Metadata ({category}): {', '.join(columns)}")

    except Exception as exc:
        print(f"⚠ Column classification failed: {exc}")
        print("Falling back to default schema...")
        schema = create_default_schema(headers)
        save_column_schema(schema, args.schema_output)
        print(f"Default schema saved to {args.schema_output}")


if __name__ == "__main__":
    main()
